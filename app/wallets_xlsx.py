# -*- coding: utf-8 -*-
"""
Чтение списка кошельков из Excel (wallets.xlsx).
Столбцы: A = приватный ключ, B = адрес депозита Bitget (куда вернуть ETH),
         C = прокси аккаунта (опц.): http://user:pass@host:port | host:port:user:pass | host:port.
Первая строка может быть заголовком (определяется автоматически и пропускается).
Пустые строки и строки, начинающиеся с '#', игнорируются.
"""

import os

import paths  # noqa: F401

DEFAULT_FILE = os.path.join(paths.CONFIG_DIR, "wallets.xlsx")
_HEADER_HINTS = ("priv", "ключ", "key", "адрес", "address", "bitget", "кошел")


def _looks_pk(s):
    t = (s or "").strip()
    if t[:2].lower() == "0x":
        t = t[2:]
    return len(t) == 64 and all(c in "0123456789abcdefABCDEF" for c in t)


def read_wallets(path=None):
    """-> список dict {private_key, bitget_address, proxy}. Бросает FileNotFoundError/ValueError."""
    from openpyxl import load_workbook
    path = path or DEFAULT_FILE
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Не найден файл со списком кошельков: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cells = [("" if c is None else str(c)).strip() for c in row]
        if not any(cells):
            continue
        pk = cells[0]
        addr = cells[1] if len(cells) > 1 else ""
        proxy = cells[2] if len(cells) > 2 else ""
        if not pk or pk.startswith("#"):
            continue
        # строка-заголовок (не приватник + похоже на подпись столбца) — пропускаем
        if not _looks_pk(pk) and any(h in pk.lower() for h in _HEADER_HINTS):
            continue
        out.append({"private_key": pk, "bitget_address": addr, "proxy": proxy})
    wb.close()
    return out


_count_cache = {}  # path -> (mtime, count): чтобы меню не перечитывало файл на каждый кадр


def count_wallets(path=None):
    """Сколько кошельков (строк с приватником) в файле. 0 — если файла нет/ошибка.
    Кэшируется по времени правки файла, поэтому частые вызовы из меню дёшевы."""
    path = path or DEFAULT_FILE
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return 0
    cached = _count_cache.get(path)
    if cached and cached[0] == mtime:
        return cached[1]
    try:
        n = len(read_wallets(path))
    except Exception:
        return cached[1] if cached else 0
    _count_cache[path] = (mtime, n)
    return n
